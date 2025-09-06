import os
import csv
import zipfile
import openpyxl
from pathlib import Path

# --- 定数定義 ---
# このスクリプトの場所を基準にプロジェクトルートを特定
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
DOWNLOAD_DIR = PROJECT_ROOT / "data" / "download"
RAW_DIR = PROJECT_ROOT / "data" / "raw"

def convert_excel_to_csv_low_memory(excel_source, file_stem, output_dir):
    """
    Excelファイルを低メモリ消費で読み込み、シートごとにCSVへ変換する。
    セル内の改行は '\\n' のような文字列にエスケープする。
    """
    try:
        # read_only=True: メモリ使用量を大幅に削減する最適化
        workbook = openpyxl.load_workbook(excel_source, read_only=True)
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # 出力ファイル名を作成 (例: database2017_H29セグメントシート.csv)
            output_filename = f"{file_stem}_{sheet_name}.csv"
            output_path = output_dir / output_filename
            
            print(f"  - Saving sheet: '{sheet_name}' -> '{output_path.name}'")
            
            # newline='': csvモジュールの公式推奨
            # encoding='utf-8-sig': Excelでの文字化けを防ぐBOM付きUTF-8
            with open(output_path, 'w', newline='', encoding='utf-8-sig') as csv_file:
                # quotingはカンマ等を含むセルを正しく扱うために依然として必要
                csv_writer = csv.writer(csv_file, quoting=csv.QUOTE_MINIMAL)

                # values_only=True: Cellオブジェクトではなくセルの値のみを取得する最適化
                for row in worksheet.iter_rows(values_only=True):
                    # 各セルを文字列に変換し、改行コードをエスケープされた文字列 '\\n' と '\\r' に置換する
                    escaped_row = []
                    for cell in row:
                        cell_str = str(cell) if cell is not None else ""
                        # Windowsの改行(\r\n)とUnix(\n),古いMac(\r)の全てに対応
                        escaped_str = cell_str.replace('\r\n', '\\n').replace('\n', '\\n').replace('\r', '\\r')
                        escaped_row.append(escaped_str)
                    
                    csv_writer.writerow(escaped_row)

    except Exception as e:
        print(f"  [Error] Failed to process {file_stem}: {e}")

def main():
    """
    downloadフォルダ内のzipとxlsxを処理し、rawフォルダにCSVを出力するメイン関数
    """
    print("--- 01_convert_to_csv.py (Low Memory & Escape Newlines): Start ---")

    RAW_DIR.mkdir(exist_ok=True)
    # (以降のmain関数は変更ありません)
    print(f"Output directory: '{RAW_DIR}'")

    source_paths = list(DOWNLOAD_DIR.glob('*.zip')) + list(DOWNLOAD_DIR.glob('*.xlsx'))
    
    if not source_paths:
        print("\n[Warning] No .zip or .xlsx files found in 'data/download/' directory.")
        print("Please download the source data first.")
        print("--- 01_convert_to_csv.py: Finished ---")
        return

    print(f"\nFound {len(source_paths)} files to process.")

    for path in source_paths:
        print(f"\nProcessing '{path.name}'...")
        
        if path.suffix == '.zip':
            try:
                with zipfile.ZipFile(path, 'r') as zf:
                    for file_in_zip in zf.namelist():
                        if file_in_zip.endswith('.xlsx'):
                            file_stem = Path(file_in_zip).stem
                            print(f"  - Found Excel file in zip: '{file_in_zip}'")
                            with zf.open(file_in_zip) as excel_stream:
                                convert_excel_to_csv_low_memory(excel_stream, file_stem, RAW_DIR)

            except zipfile.BadZipFile:
                print(f"  [Error] Bad zip file: {path.name}")
            except Exception as e:
                print(f"  [Error] Failed to process zip file {path.name}: {e}")

        elif path.suffix == '.xlsx':
            convert_excel_to_csv_low_memory(path, path.stem, RAW_DIR)

    print("\n--- 01_convert_to_csv.py: Finished ---")


if __name__ == "__main__":
    main()